#!/usr/bin/env python3
"""
=======================================================
  Sistema de Control de Inventario - Rollos
  Requisitos:  pip install textual openpyxl xlrd
=======================================================
"""
from __future__ import annotations

import fnmatch
import glob
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime

# ── Dependency checks ────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("Falta 'openpyxl'. Instálalo con:  pip install openpyxl")

try:
    import xlrd
except ImportError:
    sys.exit("Falta 'xlrd'. Instálalo con:  pip install xlrd")

try:
    from textual import on, work
    from textual.app import App, ComposeResult
    from textual.binding import Binding
    from textual.containers import Container, Horizontal, Vertical
    from textual.screen import Screen
    from textual.widgets import Button, DataTable, Footer, Header, Input, Label, Rule, Static
except ImportError:
    sys.exit("Falta 'textual'. Instálalo con:  pip install textual")

try:
    from checksum_rollo import validar_lote as _validar_lote
except ImportError:
    def _validar_lote(codigo: str) -> tuple[bool, str]:  # type: ignore[misc]
        return True, ""


# ─────────────────────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────────────────────
EXCEL_FILE        = "inventario_rollos.xlsx"
SHEET_UBICACIONES = "Ubicaciones"
SHEET_REGISTROS   = "RegistroRollos"
INVENTARIO_PREFIX = "INV_"
CREDENTIALS       = {"123": "Operador1"}


# ─────────────────────────────────────────────────────────────────────────────
# Modelos de datos
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class Ubicacion:
    deposito: str  # Nivel 1 / col F
    pasillo:  str  # Nivel 2 / col G
    columna:  str  # Nivel 3 / col H
    nivel:    str  # Nivel 4 / col I

    @property
    def id(self) -> str:
        # Concatenacion directa sin separador: ej. D0770000
        return (
            self.deposito.upper()
            + self.pasillo.upper()
            + self.columna.upper()
            + self.nivel.upper()
        )

    def to_row(self) -> list:
        return [
            self.id,
            self.deposito.upper(),
            self.pasillo.upper(),
            self.columna.upper(),
            self.nivel.upper(),
        ]


@dataclass
class RegistroRollo:
    id_usuario:   str
    id_lote:      str
    id_ubicacion: str
    fecha: str = field(
        default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )

    def to_row(self) -> list:
        return [self.id_usuario, self.id_lote, self.id_ubicacion, self.fecha]


# ─────────────────────────────────────────────────────────────────────────────
# Gestor de Excel
# ─────────────────────────────────────────────────────────────────────────────
class ExcelManager:
    def __init__(self, filepath: str = EXCEL_FILE):
        self.filepath = filepath
        self._init()

    def _init(self) -> None:
        if os.path.exists(self.filepath):
            self.wb = load_workbook(self.filepath)
        else:
            self.wb = Workbook()
            default = self.wb.active
            if default is not None:
                self.wb.remove(default)

        modified = False
        for name, headers in [
            (SHEET_UBICACIONES, ["ID", "Deposito", "Pasillo", "Columna", "Nivel"]),
            (SHEET_REGISTROS,   ["ID_Usuario", "ID_Lote", "ID_Ubicacion", "Fecha"]),
        ]:
            if name not in self.wb.sheetnames:
                ws = self.wb.create_sheet(name)
                ws.append(headers)
                self._style_header(ws)
                modified = True

        if modified:
            try:
                self.wb.save(self.filepath)
            except PermissionError:
                sys.exit(
                    f"\nERROR: No se puede guardar '{self.filepath}'.\n"
                    "El archivo esta abierto en Excel u otro programa.\n"
                    "Cerrarlo y volver a intentar.\n"
                )

    @staticmethod
    def _style_header(ws) -> None:
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def _reload(self) -> None:
        if os.path.exists(self.filepath):
            self.wb = load_workbook(self.filepath)

    # ── Lectura ──────────────────────────────────────────────────────────────
    def get_ubicaciones(self) -> list[Ubicacion]:
        self._reload()
        ws = self.wb[SHEET_UBICACIONES]
        return [
            Ubicacion(
                str(r[1] or ""),
                str(r[2] or ""),
                str(r[3] or ""),
                str(r[4] or "") if len(r) > 4 else "",
            )
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[0]
        ]

    def get_registros(self) -> list[RegistroRollo]:
        self._reload()
        ws = self.wb[SHEET_REGISTROS]
        return [
            RegistroRollo(str(r[0] or ""), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""))
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[0]
        ]

    # ── Escritura ─────────────────────────────────────────────────────────────
    def add_ubicacion(self, u: Ubicacion) -> tuple[bool, str]:
        if any(x.id == u.id for x in self.get_ubicaciones()):
            return False, f"'{u.id}' ya existe en el sistema. Usela directamente en 'Registrar Rollo'."
        ws = self.wb[SHEET_UBICACIONES]
        ws.append(u.to_row())
        self.wb.save(self.filepath)
        return True, f"Ubicacion '{u.id}' creada correctamente."

    def lote_exists(self, id_lote: str) -> bool:
        self._reload()
        ws = self.wb[SHEET_REGISTROS]
        lote = id_lote.strip()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[1] or "") == lote:
                return True
        return False

    def add_registro(self, r: RegistroRollo) -> tuple[bool, str]:
        if self.lote_exists(r.id_lote):
            return False, f"El rollo '{r.id_lote}' ya fue registrado anteriormente."
        ws = self.wb[SHEET_REGISTROS]
        ws.append(r.to_row())
        self.wb.save(self.filepath)
        return True, f"Rollo '{r.id_lote}' registrado en '{r.id_ubicacion}'."

    def get_registros_por_ubicacion(self, id_ub: str) -> list[RegistroRollo]:
        """Retorna los registros de una ubicacion especifica."""
        self._reload()
        ws  = self.wb[SHEET_REGISTROS]
        ub  = id_ub.strip().upper()
        return [
            RegistroRollo(str(r[0] or ""), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""))
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r and str(r[2] or "").upper() == ub
        ]

    def eliminar_registro(self, id_lote: str) -> tuple[bool, str]:
        """Elimina el registro de un rollo por su id_lote."""
        self._reload()
        ws   = self.wb[SHEET_REGISTROS]
        lote = id_lote.strip().upper()
        for row in ws.iter_rows(min_row=2):
            if str(row[1].value or "").upper() == lote:
                ws.delete_rows(row[1].row)
                self.wb.save(self.filepath)
                return True, f"Rollo '{lote}' eliminado correctamente."
        return False, f"Rollo '{lote}' no encontrado en el sistema."

    def ubicacion_exists(self, id_ub: str) -> bool:
        return any(u.id == id_ub.strip().upper() for u in self.get_ubicaciones())

    def import_from_estoque(self, filepath: str) -> dict:
        """
        Lee el archivo ESTOQUE .xls y carga las ubicaciones unicas.
        Columnas: F=Nivel1(deposito) G=Nivel2(pasillo) H=Nivel3(columna) I=Nivel4(nivel)
        Retorna dict con claves: total, nuevas, duplicadas, errores, archivo
        """
        result = {"total": 0, "nuevas": 0, "duplicadas": 0, "errores": 0, "archivo": filepath}

        try:
            wb_src = xlrd.open_workbook(filepath)
            ws_src = wb_src.sheet_by_index(0)
        except Exception as e:
            raise ValueError(f"No se pudo abrir el archivo: {e}")

        # Fila 0 = titulo, Fila 1 = encabezados, datos desde fila 2
        existentes = {u.id for u in self.get_ubicaciones()}
        ws_dest = self.wb[SHEET_UBICACIONES]
        nuevas_batch: list[Ubicacion] = []
        ids_vistos: set[str] = set()

        for i in range(2, ws_src.nrows):
            row = ws_src.row_values(i)
            dep = str(row[5]).strip() if row[5] else ""
            pas = str(row[6]).strip() if row[6] else ""
            col = str(row[7]).strip() if row[7] else ""
            niv = str(row[8]).strip() if row[8] else ""

            if not dep:
                continue

            result["total"] += 1
            u = Ubicacion(dep, pas, col, niv)

            if u.id in existentes or u.id in ids_vistos:
                result["duplicadas"] += 1
                continue

            ids_vistos.add(u.id)
            nuevas_batch.append(u)

        for u in nuevas_batch:
            ws_dest.append(u.to_row())
            result["nuevas"] += 1

        if nuevas_batch:
            self.wb.save(self.filepath)

        return result

    # ── Inventario mensual ────────────────────────────────────────────────────
    def get_inventarios(self) -> list[str]:
        self._reload()
        return [s for s in self.wb.sheetnames if s.startswith(INVENTARIO_PREFIX)]

    def get_ubicacion_by_id(self, id_ub: str) -> "Ubicacion | None":
        for u in self.get_ubicaciones():
            if u.id == id_ub.strip().upper():
                return u
        return None

    def crear_inventario(self, nombre: str) -> tuple[bool, str]:
        self._reload()
        if nombre in self.wb.sheetnames:
            return False, f"'{nombre}' ya existe. Se agregarán registros al inventario existente."
        ws = self.wb.create_sheet(nombre)
        ws.append([
            "ID_Ubicacion", "Deposito", "Pasillo", "Columna", "Nivel",
            "ID_Lote", "Operario", "Fecha",
        ])
        self._style_header(ws)
        self.wb.save(self.filepath)
        return True, f"Inventario '{nombre}' creado."

    def lote_en_inventario(self, hoja: str, id_lote: str) -> bool:
        self._reload()
        if hoja not in self.wb.sheetnames:
            return False
        ws = self.wb[hoja]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[5] or "") == id_lote.strip():
                return True
        return False

    def add_a_inventario(
        self, hoja: str, id_ubicacion: str, id_lote: str, operario: str
    ) -> None:
        ub = self.get_ubicacion_by_id(id_ubicacion)
        ws = self.wb[hoja]
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if ub:
            ws.append([
                ub.id, ub.deposito.upper(), ub.pasillo.upper(),
                ub.columna.upper(), ub.nivel.upper(), id_lote, operario, fecha,
            ])
        else:
            ws.append([id_ubicacion, "", "", "", "", id_lote, operario, fecha])
        self.wb.save(self.filepath)

    @staticmethod
    def buscar_estoque() -> str:
        """Busca automaticamente un archivo ESTOQUE*.xls en el directorio actual."""
        matches = glob.glob("ESTOQUE*.xls") + glob.glob("ESTOQUE*.xlsx")
        return matches[0] if matches else ""

    @staticmethod
    def buscar_export_android() -> str:
        """Busca automaticamente un archivo inventario_export_*.json en el directorio actual."""
        matches = glob.glob("inventario_export_*.json") + glob.glob("inventario_export*.json")
        return matches[0] if matches else ""

    def import_from_android(self, json_path: str) -> dict:
        """Importa datos exportados desde la app Android (inventario_android.html)."""
        import json as _json

        with open(json_path, encoding="utf-8") as f:
            data = _json.load(f)

        result = {
            "reg_nuevos": 0, "reg_dup": 0,
            "inv_nuevos": 0, "inv_dup": 0,
            "inventarios": [],
        }

        # ── Leer estado actual UNA SOLA VEZ (evita _reload() dentro del loop) ──
        self._reload()

        ws_reg = self.wb[SHEET_REGISTROS]
        lotes_reg: set[str] = {
            str(r[1] or "") for r in ws_reg.iter_rows(min_row=2, values_only=True) if r and r[1]
        }

        ws_ub = self.wb[SHEET_UBICACIONES]
        ubicaciones: dict[str, Ubicacion] = {
            str(r[0]): Ubicacion(
                str(r[1] or ""), str(r[2] or ""), str(r[3] or ""),
                str(r[4] or "") if len(r) > 4 else ""
            )
            for r in ws_ub.iter_rows(min_row=2, values_only=True) if r and r[0]
        }

        inv_lotes: dict[str, set[str]] = {
            s: {str(r[5] or "") for r in self.wb[s].iter_rows(min_row=2, values_only=True) if r and r[5]}
            for s in self.wb.sheetnames if s.startswith(INVENTARIO_PREFIX)
        }

        # ── Escribir sin recargar ─────────────────────────────────────────────
        for r in data.get("registros", []):
            lote = str(r.get("id_lote", "")).strip()
            if not lote:
                continue
            if lote in lotes_reg:
                result["reg_dup"] += 1
            else:
                lotes_reg.add(lote)
                ws_reg.append([
                    str(r.get("id_usuario", "")),
                    lote,
                    str(r.get("id_ubicacion", "")).upper(),
                    str(r.get("fecha", "")),
                ])
                result["reg_nuevos"] += 1

        for inv_nombre, registros in data.get("inventarios", {}).items():
            if inv_nombre not in self.wb.sheetnames:
                ws_i = self.wb.create_sheet(inv_nombre)
                ws_i.append([
                    "ID_Ubicacion", "Deposito", "Pasillo", "Columna", "Nivel",
                    "ID_Lote", "Operario", "Fecha",
                ])
                self._style_header(ws_i)
                inv_lotes[inv_nombre] = set()
            if inv_nombre not in result["inventarios"]:
                result["inventarios"].append(inv_nombre)

            ws_i   = self.wb[inv_nombre]
            existe = inv_lotes.setdefault(inv_nombre, set())
            fecha_imp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for r in registros:
                lote = str(r.get("id_lote", "")).strip()
                if not lote:
                    continue
                if lote in existe:
                    result["inv_dup"] += 1
                else:
                    existe.add(lote)
                    ub_id   = str(r.get("id_ubicacion", "")).upper()
                    operario = str(r.get("operario", ""))
                    fecha   = str(r.get("fecha", "")) or fecha_imp
                    ub = ubicaciones.get(ub_id)
                    if ub:
                        ws_i.append([
                            ub.id, ub.deposito.upper(), ub.pasillo.upper(),
                            ub.columna.upper(), ub.nivel.upper(),
                            lote, operario, fecha,
                        ])
                    else:
                        ws_i.append([ub_id, "", "", "", "", lote, operario, fecha])
                    result["inv_nuevos"] += 1

        self.wb.save(self.filepath)
        return result


# ─────────────────────────────────────────────────────────────────────────────
# Pantallas TUI
# ─────────────────────────────────────────────────────────────────────────────

class LoginScreen(Screen):
    """Pantalla de inicio de sesion."""

    BINDINGS = [Binding("escape", "app.quit", "Salir")]

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Container(classes="center_wrap"):
            with Vertical(classes="scan_box_narrow"):
                yield Static("CONTROL DE INVENTARIO", classes="box_title")
                yield Static("Ingrese sus credenciales para continuar", classes="box_hint")
                yield Rule()
                yield Label("Usuario:")
                yield Input(placeholder="Ingresar usuario", id="inp_login_usr")
                yield Label("Contraseña:")
                yield Input(placeholder="Ingresar contraseña", password=True, id="inp_login_pwd")
                yield Static("", id="msg_login")
                yield Rule()
                with Horizontal(classes="btn_row"):
                    yield Button("Ingresar [Enter]", id="btn_login_ok",  variant="primary")
                    yield Button("Salir [Esc]",      id="btn_login_sal", variant="error")
        yield Footer()

    def on_mount(self) -> None:
        self.query_one("#inp_login_usr", Input).focus()

    @on(Input.Submitted, "#inp_login_usr")
    def _usr_sub(self, _): self.query_one("#inp_login_pwd", Input).focus()

    @on(Input.Submitted, "#inp_login_pwd")
    def _pwd_sub(self, _): self._ingresar()

    @on(Button.Pressed, "#btn_login_ok")
    def _btn_ok(self): self._ingresar()

    @on(Button.Pressed, "#btn_login_sal")
    def _btn_sal(self): self.app.exit()

    def _ingresar(self) -> None:
        usr = self.query_one("#inp_login_usr", Input).value.strip()
        pwd = self.query_one("#inp_login_pwd", Input).value
        if not usr or not pwd:
            self._set_msg("Ingrese usuario y contraseña.", "error")
            return
        if CREDENTIALS.get(usr) != pwd:
            self._set_msg("Usuario o contraseña incorrectos.", "error")
            self.query_one("#inp_login_pwd", Input).value = ""
            self.query_one("#inp_login_pwd", Input).focus()
            return
        self.app.operario = usr
        self.app.switch_screen(MainMenuScreen())

    def _set_msg(self, text: str, kind: str) -> None:
        w = self.query_one("#msg_login", Static)
        w.update(text)
        w.remove_class("msg_ok", "msg_error")
        if kind == "ok":    w.add_class("msg_ok")
        if kind == "error": w.add_class("msg_error")


class MainMenuScreen(Screen):
    BINDINGS = [
        Binding("q", "app.quit", "Salir"),
        Binding("1", "ir_reg",   "", show=False),
        Binding("2", "ir_ub",    "", show=False),
        Binding("3", "ir_vub",   "", show=False),
        Binding("4", "ir_vreg",  "", show=False),
        Binding("5", "ir_imp",   "", show=False),
        Binding("6", "ir_inv",   "", show=False),
        Binding("7", "ir_and",   "", show=False),
    ]

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Container(classes="center_wrap"):
            with Vertical(classes="menu_box"):
                yield Static("CONTROL DE INVENTARIO", classes="box_title")
                yield Static("", id="lbl_operario", classes="operario_badge")
                yield Rule()
                yield Button("[ 1 ]  Registrar Rollo en Ubicacion", id="btn_reg",  variant="primary", classes="full_btn")
                yield Button("[ 2 ]  Nueva Ubicacion",              id="btn_ub",   variant="default", classes="full_btn")
                yield Button("[ 3 ]  Ver Ubicaciones",              id="btn_vub",  variant="default", classes="full_btn")
                yield Button("[ 4 ]  Ver Registros de Rollos",      id="btn_vreg", variant="default", classes="full_btn")
                yield Button("[ 5 ]  Importar Ubicaciones ESTOQUE", id="btn_imp",  variant="warning", classes="full_btn")
                yield Button("[ 6 ]  Nuevo Inventario Mensual",    id="btn_inv",  variant="success", classes="full_btn")
                yield Button("[ 7 ]  Importar de Tablet Android",   id="btn_and",  variant="warning", classes="full_btn")
                yield Rule()
                yield Button("[ Q ]  Salir",                        id="btn_sal",  variant="error",   classes="full_btn")
        yield Footer()

    def on_mount(self) -> None:
        self._actualizar_operario()

    def on_screen_resume(self) -> None:
        self._actualizar_operario()

    def _actualizar_operario(self) -> None:
        op = self.app.operario
        lbl = self.query_one("#lbl_operario", Static)
        if op:
            lbl.update(f"Operario activo:  {op}")
            lbl.add_class("operario_ok")
            lbl.remove_class("operario_vacio")
        else:
            lbl.update("Operario no definido  (se pedira al registrar)")
            lbl.add_class("operario_vacio")
            lbl.remove_class("operario_ok")

    @on(Button.Pressed, "#btn_reg")
    def _reg(self):   self.app.push_screen(RegistrarRolloScreen())

    @on(Button.Pressed, "#btn_ub")
    def _ub(self):    self.app.push_screen(NuevaUbicacionScreen())

    @on(Button.Pressed, "#btn_vub")
    def _vub(self):   self.app.push_screen(VerUbicacionesScreen())

    @on(Button.Pressed, "#btn_vreg")
    def _vreg(self):  self.app.push_screen(VerRegistrosScreen())

    @on(Button.Pressed, "#btn_imp")
    def _imp(self):   self.app.push_screen(ImportarEstoqueScreen())

    @on(Button.Pressed, "#btn_inv")
    def _inv(self):   self.app.push_screen(NuevoInventarioScreen())

    @on(Button.Pressed, "#btn_and")
    def _and(self):   self.app.push_screen(ImportarAndroidScreen())

    @on(Button.Pressed, "#btn_sal")
    def _sal(self):   self.app.exit()

    # ── Accesos directos teclado (teclas 1-7) ────────────────────────────────
    def action_ir_reg(self):   self.app.push_screen(RegistrarRolloScreen())
    def action_ir_ub(self):    self.app.push_screen(NuevaUbicacionScreen())
    def action_ir_vub(self):   self.app.push_screen(VerUbicacionesScreen())
    def action_ir_vreg(self):  self.app.push_screen(VerRegistrosScreen())
    def action_ir_imp(self):   self.app.push_screen(ImportarEstoqueScreen())
    def action_ir_inv(self):   self.app.push_screen(NuevoInventarioScreen())
    def action_ir_and(self):   self.app.push_screen(ImportarAndroidScreen())


class RegistrarRolloScreen(Screen):
    """
    Tres pasos con UN SOLO CAMPO por paso (optimizado para lector de barras):
      Paso 0 — scan operario  (solo si no esta definido en la sesion)
      Paso 1 — scan ubicacion
      Paso 2 — scan rollo x N  →  Finalizar y Guardar
    """

    BINDINGS = [Binding("escape", "back", "Volver / Cancelar")]

    def __init__(self):
        super().__init__()
        self._ubicacion:    str = ""
        self._pending:      list[RegistroRollo] = []
        self._lotes_sesion: set[str] = set()

    # ── Compose ───────────────────────────────────────────────────────────────
    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)

        # PASO 1 — escanear ubicacion
        with Container(id="paso1", classes="center_wrap"):
            with Vertical(classes="scan_box_narrow"):
                yield Static("ESCANEAR UBICACION", classes="box_title")
                yield Static("", id="lbl_op1", classes="operario_badge")
                yield Rule()
                yield Input(placeholder="Escanear codigo de ubicacion", id="inp_ub")
                yield Static("", id="msg1")
                yield Rule()
                with Horizontal(classes="btn_row"):
                    yield Button("Abrir [Enter]",        id="btn_abrir",   variant="primary")
                    yield Button("Volver al Menu [Esc]",  id="btn_bk1",    variant="default")

        # PASO 1b — ubicacion con rollos ya registrados
        with Container(id="paso1b", classes="scan_wrap"):
            with Horizontal(classes="ubic_banner"):
                yield Static("UBICACION:", classes="ubic_label")
                yield Static("---", id="lbl_ub1b_id", classes="ubic_id")
                yield Static("  |  OP:", classes="ubic_label")
                yield Static("---", id="lbl_ub1b_op", classes="ubic_id")
            yield Static("", id="lbl_count1b", classes="count_label")
            yield DataTable(id="tbl_ub_rollos", zebra_stripes=True)
            yield Rule()
            yield Input(placeholder="Escanear rollo a eliminar", id="inp_del")
            yield Static("", id="msg1b")
            yield Rule()
            with Horizontal(classes="btn_row"):
                yield Button("Eliminar Escaneado",      id="btn_del_rollo", variant="error")
                yield Button("Agregar mas Rollos",       id="btn_agregar",   variant="success")
                yield Button("Cambiar Ubicacion [Esc]",  id="btn_cam_ub",    variant="default")

        # PASO 2 — escanear rollos uno a uno
        with Container(id="paso2", classes="scan_wrap"):
            with Horizontal(classes="ubic_banner"):
                yield Static("UBICACION:", classes="ubic_label")
                yield Static("---", id="lbl_ubic", classes="ubic_id")
                yield Static("  |  OPERARIO:", classes="ubic_label")
                yield Static("---", id="lbl_usr2", classes="ubic_id")
            yield Rule()
            yield Input(placeholder="Escanear codigo de rollo (lote)", id="inp_lote")
            yield Static("", id="msg2")
            yield Static("Rollos en la tanda: 0", id="lbl_count", classes="count_label")
            yield DataTable(id="tbl_rollos", zebra_stripes=True)
            yield Rule()
            with Horizontal(classes="btn_row"):
                yield Button("Guardar y Continuar",   id="btn_continuar",  variant="success")
                yield Button("Guardar y Finalizar",    id="btn_finalizar",  variant="primary")
                yield Button("Cancelar Tanda [Esc]",  id="btn_cancelar",   variant="warning")

        yield Footer()

    # ── Mount ─────────────────────────────────────────────────────────────────
    def on_mount(self) -> None:
        self.query_one("#tbl_rollos",    DataTable).add_columns("#", "ID Lote", "Hora")
        self.query_one("#tbl_ub_rollos", DataTable).add_columns("ID Lote", "Operario", "Fecha")
        self._mostrar_paso(1)

    def on_input_changed(self, event: Input.Changed) -> None:
        if event.input.id in ("inp_ub", "inp_lote", "inp_del") and event.value != event.value.upper():
            pos = event.input.cursor_position
            event.input.value = event.value.upper()
            event.input.cursor_position = pos

    # ── Navegacion Escape segun paso activo ───────────────────────────────────
    def action_back(self) -> None:
        if self.query_one("#paso2").display:
            self._cancelar_tanda()
        elif self.query_one("#paso1b").display:
            self._mostrar_paso(1)
        else:
            self.app.pop_screen()

    # ─────────────────────────────────────────────────────────────────────────
    # PASO 1 — Ubicacion
    # ─────────────────────────────────────────────────────────────────────────
    @on(Input.Submitted, "#inp_ub")
    def _ub_submitted(self, _): self._abrir_ubicacion()

    @on(Button.Pressed, "#btn_abrir")
    def _btn_abrir(self): self._abrir_ubicacion()

    @on(Button.Pressed, "#btn_bk1")
    def _bk1(self): self.app.pop_screen()

    def _abrir_ubicacion(self) -> None:
        ub = self.query_one("#inp_ub", Input).value.strip().upper()
        if not ub:
            self._set_msg("#msg1", "Escanee la ubicacion.", "error")
            return
        if not ub.startswith("D"):
            self._set_msg("#msg1", "La ubicacion debe comenzar con D (DEP1, DEP2, D077, etc.).", "error")
            return
        if not self.app.excel.ubicacion_exists(ub):
            self._set_msg("#msg1", f"Ubicacion '{ub}' no existe. Creela desde el menu.", "error")
            return

        self._ubicacion = ub
        self._pending.clear()
        self._lotes_sesion.clear()

        registros = self.app.excel.get_registros_por_ubicacion(ub)
        if registros:
            self._mostrar_paso_1b(registros)
        else:
            self._ir_a_paso2()

    # ─────────────────────────────────────────────────────────────────────────
    # PASO 2 — Escaneo de rollos
    # ─────────────────────────────────────────────────────────────────────────
    @on(Input.Submitted, "#inp_lote")
    def _lote_submitted(self, _): self._agregar_lote()

    @on(Button.Pressed, "#btn_continuar")
    def _btn_continuar(self): self._guardar_y_continuar()

    @on(Button.Pressed, "#btn_finalizar")
    def _btn_finalizar(self): self._guardar_y_finalizar()

    @on(Button.Pressed, "#btn_cancelar")
    def _btn_cancelar(self): self._cancelar_tanda()

    def _agregar_lote(self) -> None:
        lote = self.query_one("#inp_lote", Input).value.strip().upper()
        if not lote:
            return

        valido, error_msg = _validar_lote(lote)
        if not valido:
            self._set_msg("#msg2", f"Codigo invalido: {error_msg}", "error")
            self.query_one("#inp_lote", Input).value = ""
            self.query_one("#inp_lote", Input).focus()
            return

        if lote in self._lotes_sesion:
            self._set_msg("#msg2", f"DUPLICADO en esta tanda: '{lote}' ya fue escaneado.", "error")
            self.query_one("#inp_lote", Input).value = ""
            self.query_one("#inp_lote", Input).focus()
            return

        self._lotes_sesion.add(lote)
        hora = datetime.now().strftime("%H:%M:%S")
        self._pending.append(RegistroRollo(self.app.operario, lote, self._ubicacion))

        n = len(self._pending)
        self.query_one("#tbl_rollos", DataTable).add_row(str(n), lote, hora)
        self.query_one("#lbl_count", Static).update(f"Rollos en la tanda: {n}")
        self.query_one("#inp_lote", Input).value = ""
        self._set_msg("#msg2", f"'{lote}' agregado  ({n} en total)", "ok")
        self.query_one("#inp_lote", Input).focus()

    # ─────────────────────────────────────────────────────────────────────────
    # PASO 1b — Ubicacion con rollos existentes
    # ─────────────────────────────────────────────────────────────────────────
    @on(Input.Submitted, "#inp_del")
    def _del_submitted(self, _): self._eliminar_rollo()

    @on(Button.Pressed, "#btn_del_rollo")
    def _btn_del(self): self._eliminar_rollo()

    @on(Button.Pressed, "#btn_agregar")
    def _btn_agregar(self): self._ir_a_paso2()

    @on(Button.Pressed, "#btn_cam_ub")
    def _btn_cam_ub(self): self._mostrar_paso(1)

    def _mostrar_paso_1b(self, registros: list) -> None:
        self.query_one("#lbl_ub1b_id", Static).update(self._ubicacion)
        self.query_one("#lbl_ub1b_op", Static).update(self.app.operario)
        self._actualizar_tabla_1b(registros)
        self.query_one("#inp_del", Input).value = ""
        self._set_msg("#msg1b", "", "")
        self._mostrar_paso("1b")

    def _actualizar_tabla_1b(self, registros: list) -> None:
        tbl = self.query_one("#tbl_ub_rollos", DataTable)
        tbl.clear()
        for r in registros:
            tbl.add_row(r.id_lote, r.id_usuario, r.fecha)
        n = len(registros)
        txt = f"{n} rollo(s) ya registrado(s)." if n else "Sin rollos — puede agregar normalmente."
        self.query_one("#lbl_count1b", Static).update(txt)

    def _eliminar_rollo(self) -> None:
        lote = self.query_one("#inp_del", Input).value.strip().upper()
        if not lote:
            self._set_msg("#msg1b", "Escanee el rollo a eliminar.", "error")
            return
        ok, msg = self.app.excel.eliminar_registro(lote)
        self._set_msg("#msg1b", msg, "ok" if ok else "error")
        if ok:
            self.query_one("#inp_del", Input).value = ""
            registros = self.app.excel.get_registros_por_ubicacion(self._ubicacion)
            self._actualizar_tabla_1b(registros)
        self.query_one("#inp_del", Input).focus()

    def _ir_a_paso2(self) -> None:
        self.query_one("#lbl_ubic",   Static).update(self._ubicacion)
        self.query_one("#lbl_usr2",   Static).update(self.app.operario)
        self.query_one("#lbl_count",  Static).update("Rollos en la tanda: 0")
        self.query_one("#tbl_rollos", DataTable).clear()
        self._set_msg("#msg2", "", "")
        self._mostrar_paso(2)

    def _guardar_y_continuar(self) -> None:
        if not self._pending:
            self._set_msg("#msg2", "No hay rollos en la tanda.", "error")
            return
        guardados, ya_existian = [], []
        for r in self._pending:
            ok, _ = self.app.excel.add_registro(r)
            (guardados if ok else ya_existian).append(r.id_lote)
        ubicacion = self._ubicacion
        self._cancelar_tanda()
        if ya_existian:
            self._set_msg("#msg1", f"{len(guardados)} guardados en {ubicacion}.  Ya existian: {len(ya_existian)}", "info")
        else:
            self._set_msg("#msg1", f"{len(guardados)} rollos guardados en {ubicacion}.", "ok")

    def _guardar_y_finalizar(self) -> None:
        for r in self._pending:
            self.app.excel.add_registro(r)
        self.app.pop_screen()

    def _cancelar_tanda(self) -> None:
        self._pending.clear()
        self._lotes_sesion.clear()
        self.query_one("#inp_ub", Input).value = ""
        self._set_msg("#msg1", "", "")
        self._mostrar_paso(1)

    # ─────────────────────────────────────────────────────────────────────────
    # Helpers
    # ─────────────────────────────────────────────────────────────────────────
    def _mostrar_paso(self, paso) -> None:
        self.query_one("#paso1").display  = (paso == 1)
        self.query_one("#paso1b").display = (paso == "1b")
        self.query_one("#paso2").display  = (paso == 2)
        if paso == 1:
            self.query_one("#lbl_op1", Static).update(f"Operario: {self.app.operario}")
            self.query_one("#inp_ub", Input).value = ""
            self.query_one("#inp_ub", Input).focus()
        elif paso == "1b":
            self.query_one("#inp_del", Input).focus()
        elif paso == 2:
            self.query_one("#inp_lote", Input).focus()

    def _set_msg(self, selector: str, text: str, kind: str) -> None:
        w = self.query_one(selector, Static)
        w.update(text)
        w.remove_class("msg_ok", "msg_error", "msg_info")
        if kind == "ok":    w.add_class("msg_ok")
        if kind == "error": w.add_class("msg_error")
        if kind == "info":  w.add_class("msg_info")


class NuevaUbicacionScreen(Screen):
    """Crea una nueva ubicacion en el deposito."""

    BINDINGS = [Binding("escape", "back", "Volver")]

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Container(classes="center_wrap"):
            with Vertical(classes="form_box"):
                yield Static("NUEVA UBICACION", classes="box_title")
                yield Static(
                    "ID = Deposito + Pasillo + Columna + Nivel  (sin separador)",
                    classes="box_hint",
                )
                yield Rule()

                yield Label("Deposito  (Nivel 1 / col F):")
                yield Input(placeholder="Ej: D077", id="inp_dep")

                yield Label("Pasillo  (Nivel 2 / col G):")
                yield Input(placeholder="Ej: 00", id="inp_pas")

                yield Label("Columna  (Nivel 3 / col H):")
                yield Input(placeholder="Ej: 00", id="inp_col")

                yield Label("Nivel  (Nivel 4 / col I):")
                yield Input(placeholder="Ej: 00", id="inp_niv")

                yield Static("ID generado: ---", id="preview", classes="id_preview")
                yield Static("", id="msg")
                yield Rule()
                with Horizontal(classes="btn_row"):
                    yield Button("Crear [Enter]", id="btn_ok", variant="success")
                    yield Button("Limpiar",       id="btn_cl", variant="default")
                    yield Button("Volver [Esc]",  id="btn_bk", variant="default")
        yield Footer()

    def on_mount(self) -> None:
        self.query_one("#inp_dep", Input).focus()

    @on(Input.Changed)
    def _preview(self, event: Input.Changed) -> None:
        if event.value != event.value.upper():
            pos = event.input.cursor_position
            event.input.value = event.value.upper()
            event.input.cursor_position = pos
            return
        d = self.query_one("#inp_dep", Input).value.strip().upper()
        p = self.query_one("#inp_pas", Input).value.strip().upper()
        c = self.query_one("#inp_col", Input).value.strip().upper()
        n = self.query_one("#inp_niv", Input).value.strip().upper()
        preview_id = (d or "?") + (p or "?") + (c or "?") + (n or "?")
        self.query_one("#preview", Static).update(f"ID generado:  {preview_id}")

    @on(Input.Submitted, "#inp_dep")
    def _next1(self, _): self.query_one("#inp_pas", Input).focus()

    @on(Input.Submitted, "#inp_pas")
    def _next2(self, _): self.query_one("#inp_col", Input).focus()

    @on(Input.Submitted, "#inp_col")
    def _next3(self, _): self.query_one("#inp_niv", Input).focus()

    @on(Input.Submitted, "#inp_niv")
    def _submit(self, _): self._crear()

    @on(Button.Pressed, "#btn_ok")
    def _btn_ok(self): self._crear()

    @on(Button.Pressed, "#btn_cl")
    def _btn_cl(self):
        for _id in ("#inp_dep", "#inp_pas", "#inp_col", "#inp_niv"):
            self.query_one(_id, Input).value = ""
        self._set_msg("", "")
        self.query_one("#inp_dep", Input).focus()

    @on(Button.Pressed, "#btn_bk")
    def action_back(self): self.app.pop_screen()

    def _crear(self) -> None:
        d = self.query_one("#inp_dep", Input).value.strip()
        p = self.query_one("#inp_pas", Input).value.strip()
        c = self.query_one("#inp_col", Input).value.strip()
        n = self.query_one("#inp_niv", Input).value.strip()

        if not d or not p or not c or not n:
            self._set_msg("Todos los campos son obligatorios.", "error")
            return
        if not d.upper().startswith("D"):
            self._set_msg("El deposito debe comenzar con D (DEP1, DEP2, D077, etc.).", "error")
            return

        ok, msg = self.app.excel.add_ubicacion(Ubicacion(d, p, c, n))
        self._set_msg(msg, "ok" if ok else "error")
        if ok:
            for _id in ("#inp_dep", "#inp_pas", "#inp_col", "#inp_niv"):
                self.query_one(_id, Input).value = ""
            self.query_one("#inp_dep", Input).focus()

    def _set_msg(self, text: str, kind: str) -> None:
        w = self.query_one("#msg", Static)
        w.update(text)
        w.remove_class("msg_ok", "msg_error")
        if kind == "ok":    w.add_class("msg_ok")
        if kind == "error": w.add_class("msg_error")


class NuevoInventarioScreen(Screen):
    """
    Inventario fisico mensual.
    Pasos: operario → nombre inventario → scan ubicacion → scan rollos → finalizar tanda
    """

    BINDINGS = [Binding("escape", "back", "Volver / Cancelar")]

    def __init__(self):
        super().__init__()
        self._inv_nombre:   str = ""
        self._ubicacion:    str = ""
        self._pending:      list[tuple[str, str]] = []   # (id_lote, hora)
        self._lotes_sesion: set[str] = set()

    # ── Compose ───────────────────────────────────────────────────────────────
    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)

        # PASO 1 — nombre del inventario
        with Container(id="inv_p1", classes="center_wrap"):
            with Vertical(classes="scan_box_narrow"):
                yield Static("INVENTARIO — ABRIR / CONTINUAR", classes="box_title")
                yield Static("", id="inv_lbl_op1", classes="operario_badge")
                yield Rule()
                yield Label("Nombre del inventario:")
                yield Input(id="inv_nombre", placeholder="Ej: INV_2026-07")
                yield Static("", id="inv_msg1")
                yield Rule()
                with Horizontal(classes="btn_row"):
                    yield Button("Abrir [Enter]",        id="inv_btn_n_ok", variant="primary")
                    yield Button("Volver al Menu [Esc]", id="inv_btn_n_bk", variant="default")

        # PASO 2 — scan ubicacion
        with Container(id="inv_p2", classes="center_wrap"):
            with Vertical(classes="scan_box_narrow"):
                yield Static("INVENTARIO — ESCANEAR UBICACION", classes="box_title")
                yield Static("", id="inv_lbl_inv", classes="inv_nombre_badge")
                yield Rule()
                yield Input(placeholder="Escanear codigo de ubicacion", id="inv_ub")
                yield Static("", id="inv_msg2")
                yield Rule()
                with Horizontal(classes="btn_row"):
                    yield Button("Abrir Ubicacion [Enter]", id="inv_btn_ub_ok", variant="primary")
                    yield Button("Cerrar Inventario",       id="inv_btn_cerrar", variant="warning")
                    yield Button("Volver al Menu [Esc]",   id="inv_btn_ub_bk",  variant="default")

        # PASO 3 — scan rollos
        with Container(id="inv_p3", classes="scan_wrap"):
            with Horizontal(classes="ubic_banner"):
                yield Static("INV:", classes="ubic_label")
                yield Static("---", id="inv_lbl_inv3",  classes="ubic_id")
                yield Static("  |  UBIC:", classes="ubic_label")
                yield Static("---", id="inv_lbl_ubic3", classes="ubic_id")
                yield Static("  |  OP:", classes="ubic_label")
                yield Static("---", id="inv_lbl_usr3",  classes="ubic_id")
            yield Rule()
            yield Input(placeholder="Escanear codigo de rollo (lote)", id="inv_lote")
            yield Static("", id="inv_msg3")
            yield Static("Rollos en la tanda: 0", id="inv_count", classes="count_label")
            yield DataTable(id="inv_tbl", zebra_stripes=True)
            yield Rule()
            with Horizontal(classes="btn_row"):
                yield Button("Guardar y Continuar",   id="inv_btn_continuar", variant="success")
                yield Button("Guardar y Finalizar",    id="inv_btn_fin",       variant="primary")
                yield Button("Cancelar Tanda [Esc]",  id="inv_btn_cancel",    variant="warning")

        yield Footer()

    # ── Mount ─────────────────────────────────────────────────────────────────
    def on_mount(self) -> None:
        self.query_one("#inv_tbl", DataTable).add_columns("#", "ID Lote", "Hora")
        nombre_auto = datetime.now().strftime(f"{INVENTARIO_PREFIX}%Y-%m")
        self.query_one("#inv_nombre", Input).value = nombre_auto
        self._mostrar_paso(1)

    def on_input_changed(self, event: Input.Changed) -> None:
        if event.input.id in ("inv_nombre", "inv_ub", "inv_lote") and event.value != event.value.upper():
            pos = event.input.cursor_position
            event.input.value = event.value.upper()
            event.input.cursor_position = pos

    # ── Escape segun paso ─────────────────────────────────────────────────────
    def action_back(self) -> None:
        if   self.query_one("#inv_p3").display: self._cancelar_tanda()
        elif self.query_one("#inv_p2").display: self.app.pop_screen()
        else: self.app.pop_screen()

    # ── PASO 1 — Nombre inventario ────────────────────────────────────────────
    @on(Input.Submitted,  "#inv_nombre")
    def _inv_nom_sub(self, _): self._abrir_inventario()
    @on(Button.Pressed,   "#inv_btn_n_ok")
    def _inv_n_ok(self):       self._abrir_inventario()
    @on(Button.Pressed,   "#inv_btn_n_bk")
    def _inv_n_bk(self):       self.app.pop_screen()

    def _abrir_inventario(self) -> None:
        nombre = self.query_one("#inv_nombre", Input).value.strip()
        if not nombre:
            self._msg("#inv_msg1", "Ingrese un nombre para el inventario.", "error")
            return
        if not nombre.startswith(INVENTARIO_PREFIX):
            nombre = INVENTARIO_PREFIX + nombre

        ok, msg = self.app.excel.crear_inventario(nombre)
        self._inv_nombre = nombre
        self._msg("#inv_msg1", msg, "ok" if ok else "info")
        self._mostrar_paso(2)

    # ── PASO 2 — Scan ubicacion ───────────────────────────────────────────────
    @on(Input.Submitted,  "#inv_ub")
    def _inv_ub_sub(self, _): self._abrir_ubicacion()
    @on(Button.Pressed,   "#inv_btn_ub_ok")
    def _inv_ub_ok(self):     self._abrir_ubicacion()
    @on(Button.Pressed,   "#inv_btn_cerrar")
    def _inv_cerrar(self):    self.app.pop_screen()
    @on(Button.Pressed,   "#inv_btn_ub_bk")
    def _inv_ub_bk(self):    self.app.pop_screen()

    def _abrir_ubicacion(self) -> None:
        ub = self.query_one("#inv_ub", Input).value.strip().upper()
        if not ub:
            self._msg("#inv_msg2", "Escanee la ubicacion.", "error")
            return
        if not ub.startswith("D"):
            self._msg("#inv_msg2", "La ubicacion debe comenzar con D (DEP1, DEP2, D077, etc.).", "error")
            return
        if not self.app.excel.ubicacion_exists(ub):
            self._msg("#inv_msg2", f"Ubicacion '{ub}' no existe.", "error")
            return

        self._ubicacion = ub
        self._pending.clear()
        self._lotes_sesion.clear()

        self.query_one("#inv_lbl_inv3",  Static).update(self._inv_nombre)
        self.query_one("#inv_lbl_ubic3", Static).update(ub)
        self.query_one("#inv_lbl_usr3",  Static).update(self.app.operario)
        self.query_one("#inv_count",     Static).update("Rollos en la tanda: 0")
        self.query_one("#inv_tbl",       DataTable).clear()
        self._msg("#inv_msg3", "", "")
        self._mostrar_paso(3)

    # ── PASO 3 — Scan rollos ──────────────────────────────────────────────────
    @on(Input.Submitted,  "#inv_lote")
    def _inv_lote_sub(self, _): self._agregar_lote()
    @on(Button.Pressed,   "#inv_btn_continuar")
    def _inv_continuar(self):   self._guardar_y_continuar()
    @on(Button.Pressed,   "#inv_btn_fin")
    def _inv_fin(self):         self._guardar_y_finalizar()
    @on(Button.Pressed,   "#inv_btn_cancel")
    def _inv_cancel(self):      self._cancelar_tanda()

    def _agregar_lote(self) -> None:
        lote = self.query_one("#inv_lote", Input).value.strip().upper()
        if not lote:
            return

        valido, error_msg = _validar_lote(lote)
        if not valido:
            self._msg("#inv_msg3", f"Codigo invalido: {error_msg}", "error")
            self.query_one("#inv_lote", Input).value = ""
            self.query_one("#inv_lote", Input).focus()
            return

        # Duplicado en esta tanda
        if lote in self._lotes_sesion:
            self._msg("#inv_msg3", f"DUPLICADO en tanda: '{lote}' ya escaneado.", "error")
            self.query_one("#inv_lote", Input).value = ""
            self.query_one("#inv_lote", Input).focus()
            return

        # Duplicado en el inventario completo
        if self.app.excel.lote_en_inventario(self._inv_nombre, lote):
            self._msg("#inv_msg3", f"'{lote}' ya figura en este inventario (otra ubicacion).", "error")
            self.query_one("#inv_lote", Input).value = ""
            self.query_one("#inv_lote", Input).focus()
            return

        self._lotes_sesion.add(lote)
        hora = datetime.now().strftime("%H:%M:%S")
        self._pending.append((lote, hora))

        n = len(self._pending)
        self.query_one("#inv_tbl",   DataTable).add_row(str(n), lote, hora)
        self.query_one("#inv_count", Static).update(f"Rollos en la tanda: {n}")
        self.query_one("#inv_lote",  Input).value = ""
        self._msg("#inv_msg3", f"'{lote}' agregado  ({n} en total)", "ok")
        self.query_one("#inv_lote", Input).focus()

    def _guardar_y_continuar(self) -> None:
        if not self._pending:
            self._msg("#inv_msg3", "No hay rollos en la tanda.", "error")
            return
        for lote, _ in self._pending:
            self.app.excel.add_a_inventario(
                self._inv_nombre, self._ubicacion, lote, self.app.operario
            )
        n = len(self._pending)
        ubicacion = self._ubicacion
        self._cancelar_tanda()
        self._msg("#inv_msg2", f"{n} rollos guardados en {ubicacion}.", "ok")

    def _guardar_y_finalizar(self) -> None:
        for lote, _ in self._pending:
            self.app.excel.add_a_inventario(
                self._inv_nombre, self._ubicacion, lote, self.app.operario
            )
        self.app.pop_screen()

    def _cancelar_tanda(self) -> None:
        self._pending.clear()
        self._lotes_sesion.clear()
        self.query_one("#inv_ub",  Input).value = ""
        self._mostrar_paso(2)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _mostrar_paso(self, paso: int) -> None:
        for pid, n in [("#inv_p1", 1), ("#inv_p2", 2), ("#inv_p3", 3)]:
            self.query_one(pid).display = (paso == n)
        if paso == 1:
            self.query_one("#inv_lbl_op1", Static).update(f"Operario: {self.app.operario}")
            self.query_one("#inv_nombre", Input).focus()
        elif paso == 2:
            self.query_one("#inv_lbl_inv", Static).update(self._inv_nombre)
            self.query_one("#inv_ub",  Input).value = ""
            self.query_one("#inv_ub",  Input).focus()
        elif paso == 3:
            self.query_one("#inv_lote", Input).focus()

    def _msg(self, selector: str, text: str, kind: str) -> None:
        w = self.query_one(selector, Static)
        w.update(text)
        w.remove_class("msg_ok", "msg_error", "msg_info")
        if kind == "ok":    w.add_class("msg_ok")
        if kind == "error": w.add_class("msg_error")
        if kind == "info":  w.add_class("msg_info")


class ImportarEstoqueScreen(Screen):
    """Importa ubicaciones unicas desde un archivo ESTOQUE .xls"""

    BINDINGS = [
        Binding("escape", "back",     "Volver"),
        Binding("f",      "explorar", "Explorar"),
    ]

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Container(classes="center_wrap"):
            with Vertical(classes="form_box"):
                yield Static("IMPORTAR UBICACIONES DEL ESTOQUE", classes="box_title")
                yield Static(
                    "Lee Nivel 1+2+3+4 (cols F-G-H-I) y carga ubicaciones unicas.",
                    classes="box_hint",
                )
                yield Rule()

                yield Label("Archivo ESTOQUE (.xls / .xlsx):")
                yield Input(id="inp_path", placeholder="Nombre o ruta del archivo")

                yield Static("", id="msg")

                with Vertical(id="resultado", classes="resultado_box"):
                    yield Static("", id="res_total")
                    yield Static("", id="res_nuevas")
                    yield Static("", id="res_dup")

                yield Rule()
                with Horizontal(classes="btn_row"):
                    yield Button("Importar",     id="btn_ok",  variant="success")
                    yield Button("Explorar [F]", id="btn_exp", variant="default")
                    yield Button("Volver [Esc]", id="btn_bk",  variant="default")
        yield Footer()

    def on_mount(self) -> None:
        archivo = self.app.excel.buscar_estoque()
        if archivo:
            self.query_one("#inp_path", Input).value = archivo
        self.query_one("#inp_path", Input).focus()
        self.query_one("#resultado").display = False

    @on(Input.Submitted, "#inp_path")
    def _submit(self, _): self._importar()

    @on(Button.Pressed, "#btn_ok")
    def _btn_ok(self): self._importar()

    @on(Button.Pressed, "#btn_exp")
    def _btn_exp(self): self.action_explorar()

    @on(Button.Pressed, "#btn_bk")
    def action_back(self): self.app.pop_screen()

    def action_explorar(self) -> None:
        def _set(path: str | None) -> None:
            if path:
                self.query_one("#inp_path", Input).value = path
        self.app.push_screen(
            FilePickerScreen("*.xls*", "Seleccionar archivo ESTOQUE (.xls / .xlsx)"), _set
        )

    def _importar(self) -> None:
        filepath = self.query_one("#inp_path", Input).value.strip()
        if not filepath:
            self._set_msg("Ingrese el nombre del archivo.", "error")
            return
        if not os.path.exists(filepath):
            self._set_msg(f"Archivo no encontrado: {filepath}", "error")
            return

        self._set_msg("Importando, espere...", "info")
        self.query_one("#resultado").display = False

        try:
            r = self.app.excel.import_from_estoque(filepath)
        except ValueError as e:
            self._set_msg(str(e), "error")
            return

        self.query_one("#res_total",  Static).update(f"  Filas leidas en ESTOQUE : {r['total']}")
        self.query_one("#res_nuevas", Static).update(f"  Ubicaciones importadas   : {r['nuevas']}")
        self.query_one("#res_dup",    Static).update(f"  Ya existian (omitidas)   : {r['duplicadas']}")
        self.query_one("#resultado").display = True

        if r["nuevas"] > 0:
            self._set_msg(f"Importacion completada: {r['nuevas']} ubicaciones nuevas agregadas.", "ok")
        else:
            self._set_msg("Todas las ubicaciones del archivo ya existian.", "info")

    def _set_msg(self, text: str, kind: str) -> None:
        w = self.query_one("#msg", Static)
        w.update(text)
        w.remove_class("msg_ok", "msg_error", "msg_info")
        if kind == "ok":    w.add_class("msg_ok")
        if kind == "error": w.add_class("msg_error")
        if kind == "info":  w.add_class("msg_info")


class ImportarAndroidScreen(Screen):
    """Importa el JSON exportado por inventario_android.html."""

    BINDINGS = [
        Binding("escape", "back",     "Volver"),
        Binding("f",      "explorar", "Explorar"),
    ]

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Container(classes="center_wrap"):
            with Vertical(classes="form_box"):
                yield Static("IMPORTAR DATOS DE TABLET ANDROID", classes="box_title")
                yield Static(
                    "Seleccione el archivo JSON exportado desde la app HTML.",
                    classes="box_hint",
                )
                yield Rule()

                yield Label("Archivo JSON (inventario_export_*.json):")
                yield Input(id="inp_json", placeholder="Nombre o ruta del archivo JSON")

                yield Static("", id="msg_and")

                with Vertical(id="res_and", classes="resultado_box"):
                    yield Static("", id="and_reg_n")
                    yield Static("", id="and_reg_d")
                    yield Static("", id="and_inv_n")
                    yield Static("", id="and_inv_d")
                    yield Static("", id="and_inv_lista")

                yield Rule()
                with Horizontal(classes="btn_row"):
                    yield Button("Importar",     id="btn_and_ok",  variant="success")
                    yield Button("Explorar [F]", id="btn_and_exp", variant="default")
                    yield Button("Volver [Esc]", id="btn_and_bk",  variant="default")
        yield Footer()

    def on_mount(self) -> None:
        archivo = self.app.excel.buscar_export_android()
        if archivo:
            self.query_one("#inp_json", Input).value = archivo
        self.query_one("#inp_json", Input).focus()
        self.query_one("#res_and").display = False

    @on(Input.Submitted, "#inp_json")
    def _submit(self, _): self._importar()

    @on(Button.Pressed, "#btn_and_ok")
    def _btn_ok(self): self._importar()

    @on(Button.Pressed, "#btn_and_exp")
    def _btn_exp(self): self.action_explorar()

    @on(Button.Pressed, "#btn_and_bk")
    def action_back(self): self.app.pop_screen()

    def action_explorar(self) -> None:
        def _set(path: str | None) -> None:
            if path:
                self.query_one("#inp_json", Input).value = path
        self.app.push_screen(
            FilePickerScreen("*.json", "Seleccionar archivo JSON exportado"), _set
        )

    def _importar(self) -> None:
        filepath = self.query_one("#inp_json", Input).value.strip()
        if not filepath:
            self._msg("Ingrese el nombre del archivo JSON.", "error")
            return
        if not os.path.exists(filepath):
            self._msg(f"Archivo no encontrado: {filepath}", "error")
            return

        self._msg("Importando, espere...", "info")
        self.query_one("#res_and").display = False

        try:
            r = self.app.excel.import_from_android(filepath)
        except Exception as e:
            self._msg(f"Error al leer el archivo: {e}", "error")
            return

        self.query_one("#and_reg_n",    Static).update(f"  Registros de rollos nuevos  : {r['reg_nuevos']}")
        self.query_one("#and_reg_d",    Static).update(f"  Rollos ya existian (omitidos): {r['reg_dup']}")
        self.query_one("#and_inv_n",    Static).update(f"  Filas de inventario nuevas   : {r['inv_nuevos']}")
        self.query_one("#and_inv_d",    Static).update(f"  Filas inv. ya existian       : {r['inv_dup']}")
        inv_lista = ", ".join(r["inventarios"]) if r["inventarios"] else "ninguno"
        self.query_one("#and_inv_lista", Static).update(f"  Inventarios procesados       : {inv_lista}")
        self.query_one("#res_and").display = True

        total = r["reg_nuevos"] + r["inv_nuevos"]
        if total > 0:
            self._msg(f"Importacion OK: {r['reg_nuevos']} rollos + {r['inv_nuevos']} filas de inventario.", "ok")
        else:
            self._msg("No habia datos nuevos. Todo ya existia.", "info")

    def _msg(self, text: str, kind: str) -> None:
        w = self.query_one("#msg_and", Static)
        w.update(text)
        w.remove_class("msg_ok", "msg_error", "msg_info")
        if kind == "ok":    w.add_class("msg_ok")
        if kind == "error": w.add_class("msg_error")
        if kind == "info":  w.add_class("msg_info")


class VerUbicacionesScreen(Screen):
    BINDINGS = [
        Binding("escape", "back", "Volver"),
        Binding("r",      "refresh", "Actualizar"),
    ]

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Container(classes="table_wrap"):
            yield Static("UBICACIONES REGISTRADAS", id="tbl_title", classes="box_title")
            yield DataTable(id="tbl", zebra_stripes=True)
            with Horizontal(classes="btn_row"):
                yield Button("Actualizar [R]", id="btn_ref", variant="default")
                yield Button("Volver [Esc]",   id="btn_bk",  variant="default")
        yield Footer()

    def on_mount(self) -> None:
        self._load()

    def _load(self) -> None:
        tbl = self.query_one("#tbl", DataTable)
        tbl.clear(columns=True)
        tbl.add_columns("ID", "Deposito", "Pasillo", "Columna", "Nivel")
        rows = self.app.excel.get_ubicaciones()
        for u in rows:
            tbl.add_row(u.id, u.deposito.upper(), u.pasillo.upper(), u.columna.upper(), u.nivel.upper())
        self.query_one("#tbl_title", Static).update(
            f"UBICACIONES REGISTRADAS  ({len(rows)})"
        )

    def action_refresh(self): self._load()

    @on(Button.Pressed, "#btn_ref")
    def _ref(self): self._load()

    @on(Button.Pressed, "#btn_bk")
    def action_back(self): self.app.pop_screen()


class VerRegistrosScreen(Screen):
    BINDINGS = [
        Binding("escape", "back", "Volver"),
        Binding("r",      "refresh", "Actualizar"),
    ]

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Container(classes="table_wrap"):
            yield Static("REGISTROS DE ROLLOS", id="tbl_title", classes="box_title")
            yield DataTable(id="tbl", zebra_stripes=True)
            with Horizontal(classes="btn_row"):
                yield Button("Actualizar [R]", id="btn_ref", variant="default")
                yield Button("Volver [Esc]",   id="btn_bk",  variant="default")
        yield Footer()

    def on_mount(self) -> None:
        self._load()

    def _load(self) -> None:
        tbl = self.query_one("#tbl", DataTable)
        tbl.clear(columns=True)
        tbl.add_columns("ID Usuario", "ID Lote", "ID Ubicacion", "Fecha")
        rows = self.app.excel.get_registros()
        for r in rows:
            tbl.add_row(r.id_usuario, r.id_lote, r.id_ubicacion, r.fecha)
        self.query_one("#tbl_title", Static).update(
            f"REGISTROS DE ROLLOS  ({len(rows)})"
        )

    def action_refresh(self): self._load()

    @on(Button.Pressed, "#btn_ref")
    def _ref(self): self._load()

    @on(Button.Pressed, "#btn_bk")
    def action_back(self): self.app.pop_screen()


# ─────────────────────────────────────────────────────────────────────────────
# Explorador de archivos
# ─────────────────────────────────────────────────────────────────────────────
class FilePickerScreen(Screen):
    """
    Explorador de archivos para seleccionar una ruta.
    Uso:  push_screen(FilePickerScreen("*.json", "Titulo"), callback)
    El callback recibe str (ruta seleccionada) o None si se cancela.
    """

    BINDINGS = [
        Binding("escape", "cancelar", "Cancelar"),
        Binding("u",      "subir",    "Subir directorio"),
    ]

    def __init__(self, patron: str = "*", titulo: str = "Seleccionar archivo"):
        super().__init__()
        self._patron   = patron
        self._titulo   = titulo
        self._dir      = os.path.abspath(os.getcwd())
        self._entradas: list[tuple[str, str]] = []  # (tipo, ruta_absoluta)

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Container(classes="picker_wrap"):
            yield Static("", id="fp_titulo", classes="box_title")
            yield Static("", id="fp_ruta",   classes="fp_ruta")
            yield Rule()
            yield DataTable(id="fp_tbl", zebra_stripes=True)
            yield Static("", id="fp_msg", classes="fp_msg")
            yield Rule()
            with Horizontal(classes="btn_row"):
                yield Button("Seleccionar [Enter]", id="fp_ok",     variant="success")
                yield Button("Subir dir.  [U]",      id="fp_up",     variant="default")
                yield Button("Cancelar    [Esc]",    id="fp_cancel", variant="error")
        yield Footer()

    def on_mount(self) -> None:
        self.query_one("#fp_titulo", Static).update(self._titulo)
        tbl = self.query_one("#fp_tbl", DataTable)
        tbl.add_columns("Tipo", "Nombre", "Tamaño", "Modificado")
        tbl.cursor_type = "row"
        self._cargar()

    # ── Carga del directorio ──────────────────────────────────────────────────
    def _cargar(self) -> None:
        tbl = self.query_one("#fp_tbl", DataTable)
        tbl.clear()
        self._entradas.clear()
        self.query_one("#fp_ruta", Static).update(self._dir)

        try:
            items = list(os.scandir(self._dir))
        except PermissionError:
            self.query_one("#fp_msg", Static).update("Sin permiso para acceder a este directorio.")
            return

        dirs  = sorted([e for e in items if e.is_dir()], key=lambda e: e.name.lower())
        files = sorted(
            [e for e in items
             if e.is_file() and fnmatch.fnmatch(e.name.lower(), self._patron.lower())],
            key=lambda e: e.name.lower(),
        )

        for e in dirs:
            tbl.add_row("[DIR]", e.name, "", "")
            self._entradas.append(("dir", e.path))

        for e in files:
            st  = e.stat()
            kb  = f"{st.st_size // 1024} KB" if st.st_size >= 1024 else f"{st.st_size} B"
            mod = datetime.fromtimestamp(st.st_mtime).strftime("%d/%m/%y %H:%M")
            tbl.add_row("[ARQ]", e.name, kb, mod)
            self._entradas.append(("file", e.path))

        msg = self.query_one("#fp_msg", Static)
        if not files and not dirs:
            msg.update("Directorio vacío.")
        elif not files:
            msg.update(f"No hay archivos '{self._patron}' aquí — navegue directorios.")
        else:
            msg.update(f"{len(files)} archivo(s)  |  {len(dirs)} carpeta(s)")
            tbl.focus()

    # ── Acción al seleccionar fila ────────────────────────────────────────────
    def _seleccionar_fila(self) -> None:
        tbl = self.query_one("#fp_tbl", DataTable)
        idx = tbl.cursor_row
        if idx < 0 or idx >= len(self._entradas):
            return
        tipo, ruta = self._entradas[idx]
        if tipo == "dir":
            self._dir = os.path.abspath(ruta)
            self._cargar()
        else:
            self.dismiss(ruta)

    def _subir_directorio(self) -> None:
        padre = os.path.dirname(self._dir)
        if padre and padre != self._dir:
            self._dir = padre
            self._cargar()

    # ── Handlers ─────────────────────────────────────────────────────────────
    @on(Button.Pressed, "#fp_ok")
    def _ok(self): self._seleccionar_fila()

    @on(Button.Pressed, "#fp_up")
    def _up(self): self._subir_directorio()

    @on(Button.Pressed, "#fp_cancel")
    def _cancel(self): self.dismiss(None)

    @on(DataTable.RowSelected)
    def _row_selected(self, _): self._seleccionar_fila()

    def action_cancelar(self): self.dismiss(None)
    def action_subir(self):    self._subir_directorio()


# ─────────────────────────────────────────────────────────────────────────────
# Aplicacion principal
# ─────────────────────────────────────────────────────────────────────────────
class InventarioApp(App):
    TITLE     = "Control de Inventario — Rollos"
    SUB_TITLE = f"Datos: {EXCEL_FILE}"

    CSS = """
    /* ── Layout ─────────────────────────────────────────────────── */
    .center_wrap {
        align: center middle;
        width: 100%;
        height: 1fr;
        overflow-y: auto;
    }
    .table_wrap {
        width: 100%;
        height: 1fr;
        padding: 1 2;
    }

    /* ── Contenedores de menu y formulario ───────────────────────── */
    .menu_box {
        width: 56;
        border: double $primary;
        background: $surface;
        padding: 1 3;
    }
    .form_box {
        width: 66;
        border: double $primary;
        background: $surface;
        padding: 1 3;
    }

    /* ── Tipografia ───────────────────────────────────────────────── */
    .box_title {
        text-align: center;
        color: $primary;
        text-style: bold;
        padding-bottom: 1;
    }
    .box_subtitle {
        text-align: center;
        color: $text-muted;
        margin-bottom: 1;
    }
    .box_hint {
        text-align: center;
        color: $text-muted;
        margin-bottom: 1;
    }
    .id_preview {
        margin-top: 1;
        color: $accent;
        text-style: bold;
    }

    /* ── Botones ──────────────────────────────────────────────────── */
    .full_btn {
        width: 100%;
        margin: 0;
        min-height: 1;
    }
    .btn_row {
        margin-top: 1;
        align: center middle;
    }
    .btn_row Button {
        margin: 0 1;
    }

    /* ── Inputs y labels ─────────────────────────────────────────── */
    Label {
        margin-top: 1;
        color: $text-muted;
    }
    Input {
        margin-bottom: 0;
    }

    /* ── Mensajes de estado ───────────────────────────────────────── */
    .msg_ok {
        color: $success;
        text-align: center;
        margin-top: 1;
        text-style: bold;
    }
    .msg_error {
        color: $error;
        text-align: center;
        margin-top: 1;
        text-style: bold;
    }
    .msg_info {
        color: $warning;
        text-align: center;
        margin-top: 1;
        text-style: bold;
    }

    /* ── Resultado de importacion ─────────────────────────────────── */
    .resultado_box {
        border: solid $primary;
        padding: 0 1;
        margin-top: 1;
        background: $surface-darken-1;
    }
    #res_total  { color: $text; }
    #res_nuevas { color: $success; text-style: bold; }
    #res_dup    { color: $text-muted; }

    /* ── Operario en menu principal ─────────────────────────────────── */
    .operario_badge {
        text-align: center;
        margin-bottom: 1;
        text-style: bold;
    }
    .operario_ok    { color: $success; }
    .operario_vacio { color: $text-muted; }

    /* ── Inventario mensual ──────────────────────────────────────────── */
    .inv_nombre_badge {
        text-align: center;
        color: $accent;
        text-style: bold;
        margin-bottom: 1;
    }

    /* ── Caja angosta para pasos de scan ─────────────────────────────── */
    .scan_box_narrow {
        width: 58;
        border: double $primary;
        background: $surface;
        padding: 1 3;
    }

    /* ── Pantalla de escaneo (paso 2) ────────────────────────────────── */
    .scan_wrap {
        width: 100%;
        height: 1fr;
        padding: 0 2 1 2;
    }
    .ubic_banner {
        background: $primary;
        color: $background;
        padding: 0 2;
        height: 3;
        margin-bottom: 1;
        align: left middle;
    }
    .ubic_label {
        color: $background;
        margin-right: 1;
    }
    .ubic_id {
        color: $accent;
        text-style: bold;
    }
    .count_label {
        color: $warning;
        text-style: bold;
        margin-top: 1;
        margin-bottom: 1;
    }

    /* ── Tabla ────────────────────────────────────────────────────── */
    DataTable {
        height: 1fr;
        margin-bottom: 1;
    }
    #tbl_ub_rollos {
        height: 12;
    }

    /* ── Explorador de archivos ───────────────────────────────────── */
    .picker_wrap {
        width: 100%;
        height: 1fr;
        padding: 1 2;
    }
    .fp_ruta {
        color: $accent;
        text-style: bold;
        margin-bottom: 1;
        padding: 0 1;
    }
    .fp_msg {
        color: $text-muted;
        margin-top: 1;
    }
    """

    def on_mount(self) -> None:
        self.operario: str = ""
        self.excel = ExcelManager()
        self.push_screen(LoginScreen())


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    InventarioApp().run()
